"""
Скачивание и сжатие изображений из публичных ссылок Яндекс.Диска
для товаров из matched_with_links.xlsx.

Ожидаемые колонки во входном Excel:
- "Артикул [ARTIKUL]"
- "Наименование элемента"
- "Ссылка на карточку"

Логика:
1. preview:
   - файл *.pdf.png в корне
   - если нет -> первая валидная картинка
2. gallery:
   - если есть папка Images (на любой глубине) -> imgNN.*
   - если нет -> все картинки из корня
3. На выходе:
   - папка по артикулу
   - preview.webp
   - gallery_01.webp ... gallery_05.webp
   - report.xlsx со статусами
"""

import os
import re
import time
from io import BytesIO
from typing import Optional, Tuple, List

import pandas as pd
import requests
from openpyxl import load_workbook
from PIL import Image

# =========================
# НАСТРОЙКИ
# =========================

INPUT_FILE = "matched_with_links.xlsx"
INPUT_SHEET = 0

OUTPUT_DIR = "upload/import_images"
REPORT_FILE = os.path.join(OUTPUT_DIR, "report.xlsx")

COL_ARTICLE = "Артикул [ARTIKUL]"
COL_NAME = "Наименование элемента"
COL_LINK = "Ссылка на карточку"

YANDEX_API = "https://cloud-api.yandex.net/v1/disk/public/resources"

MAX_GALLERY = 5
TIMEOUT = 20
SLEEP_BETWEEN_ITEMS = 0.25
MAX_DEPTH = 6

MIN_BYTES = 5000
DOWNLOAD_RETRIES = 2
RETRY_SLEEP = 0.5

WEBP_QUALITY = 80
WEBP_METHOD = 6
MAX_WIDTH = 1600
MAX_HEIGHT = 1600

# Иногда Pillow ругается на огромные изображения.
Image.MAX_IMAGE_PIXELS = None

# =========================
# ХЕЛПЕРЫ
# =========================

def is_empty(value) -> bool:
    if pd.isna(value):
        return True
    text = str(value).strip()
    return text == "" or text.lower() == "nan"


def safe_name(value: object) -> str:
    text = str(value).strip()
    text = re.sub(r"[^\w\-.]+", "_", text)
    return text[:120]


def api_items(public_key: str, path: str = "", session: Optional[requests.Session] = None) -> List[dict]:
    params = {
        "public_key": public_key,
        "path": path,
        "limit": 1000,
    }

    try:
        response = (session or requests).get(YANDEX_API, params=params, timeout=TIMEOUT)
        if response.status_code != 200:
            return []

        data = response.json()
        return data.get("_embedded", {}).get("items", []) or []
    except Exception:
        return []


def is_gallery_image(filename: str) -> bool:
    return bool(re.match(r"^img\d+\.(jpg|jpeg|png|webp)$", filename.lower()))


def is_any_image(filename: str) -> bool:
    return bool(re.search(r"\.(jpg|jpeg|png|webp)$", filename.lower()))


def is_preview_image(filename: str) -> bool:
    return filename.lower().endswith(".pdf.png")


def prepare_image_for_webp(image: Image.Image) -> Image.Image:
    """
    Готовим изображение для webp:
    - сохраняем прозрачность, если она есть
    - приводим режим к RGB/RGBA
    - уменьшаем слишком большие размеры
    """
    has_alpha = image.mode in ("RGBA", "LA") or (image.mode == "P" and "transparency" in image.info)

    if has_alpha:
        image = image.convert("RGBA")
    else:
        image = image.convert("RGB")

    image.thumbnail((MAX_WIDTH, MAX_HEIGHT), Image.LANCZOS)
    return image


def download_and_compress_image(url: str, target_path: str, session: requests.Session) -> Tuple[bool, str]:
    last_reason = "unknown"

    for _ in range(DOWNLOAD_RETRIES + 1):
        try:
            response = session.get(url, timeout=TIMEOUT)
            if response.status_code != 200:
                last_reason = f"http_{response.status_code}"
                raise Exception()

            content_type = (response.headers.get("Content-Type") or "").lower()
            if "text/html" in content_type:
                last_reason = "html_response"
                raise Exception()

            if len(response.content) < MIN_BYTES:
                last_reason = "too_small"
                raise Exception()

            image = Image.open(BytesIO(response.content))
            image = prepare_image_for_webp(image)

            image.save(
                target_path,
                format="WEBP",
                quality=WEBP_QUALITY,
                method=WEBP_METHOD,
            )

            return True, "ok"

        except Exception as e:
            text = str(e).strip()
            if text:
                last_reason = text
            time.sleep(RETRY_SLEEP)

    return False, last_reason


def find_images_dir(public_key: str, session: requests.Session) -> Tuple[Optional[str], str]:
    root_items = api_items(public_key, "", session)
    queue = [(item["path"], 1) for item in root_items if item.get("type") == "dir"]

    while queue:
        current_path, depth = queue.pop(0)
        items = api_items(public_key, current_path, session)

        for item in items:
            if item.get("type") == "dir" and item.get("name", "").lower() == "images":
                return item["path"], f"папка Images найдена на глубине {depth + 1}"

        if depth < MAX_DEPTH:
            for item in items:
                if item.get("type") == "dir":
                    queue.append((item["path"], depth + 1))

    return None, "папка Images не найдена"


def read_excel_with_hyperlinks(filepath: str, sheet=0) -> pd.DataFrame:
    """
    Читаем Excel так, чтобы в колонке со ссылкой был реальный URL,
    а не отображаемый текст типа 'Ссылка'.
    """
    df = pd.read_excel(filepath, sheet_name=sheet)

    wb = load_workbook(filepath, data_only=False)
    ws = wb[wb.sheetnames[sheet] if isinstance(sheet, int) else sheet]

    headers = [cell.value for cell in ws[1]]
    if COL_LINK not in headers:
        raise ValueError(f"Во входном файле нет колонки: {COL_LINK}")

    link_col_idx = headers.index(COL_LINK) + 1  # 1-based

    real_links = []
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=link_col_idx)

        if cell.hyperlink and cell.hyperlink.target:
            real_links.append(cell.hyperlink.target)
        else:
            real_links.append(cell.value)

    if len(real_links) != len(df):
        raise ValueError(
            f"Не сошлось количество строк: hyperlinks={len(real_links)}, dataframe={len(df)}"
        )

    df[COL_LINK] = real_links
    return df


# =========================
# ОСНОВНАЯ ЛОГИКА
# =========================

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    df = read_excel_with_hyperlinks(INPUT_FILE, INPUT_SHEET)

    required_columns = [COL_ARTICLE, COL_NAME, COL_LINK]
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"Во входном файле не хватает колонок: {missing_columns}")

    print("\nПроверка первых 5 ссылок из Excel:")
    for _, row in df.head(5).iterrows():
        print(f"- {row.get(COL_ARTICLE)} -> {row.get(COL_LINK)}")

    report_rows = []

    ok_count = 0
    fail_count = 0
    skip_count = 0
    partial_count = 0

    failed_articles = []
    partial_articles = []

    with requests.Session() as session:
        session.headers["User-Agent"] = "MakitaImageLoader/3.0"

        total = len(df)

        for index, row in df.iterrows():
            article = str(row.get(COL_ARTICLE, "")).strip()
            name = str(row.get(COL_NAME, "")).strip()
            public_link = str(row.get(COL_LINK, "")).strip()

            print(f"\n[{index + 1}/{total}] Обработка: {article} | {name}")

            if is_empty(article):
                skip_count += 1
                print("  -> Пропуск: пустой артикул")

                report_rows.append({
                    "Артикул": article,
                    "Наименование": name,
                    "Ссылка": public_link,
                    "Статус": "SKIP",
                    "Превью": "нет",
                    "Галерея": 0,
                    "Источник": "",
                    "Примечание": "пустой артикул",
                })
                continue

            if is_empty(public_link):
                skip_count += 1
                print("  -> Пропуск: пустая ссылка")

                report_rows.append({
                    "Артикул": article,
                    "Наименование": name,
                    "Ссылка": public_link,
                    "Статус": "SKIP",
                    "Превью": "нет",
                    "Галерея": 0,
                    "Источник": "",
                    "Примечание": "пустая ссылка",
                })
                continue

            item_dir = os.path.join(OUTPUT_DIR, safe_name(article))
            os.makedirs(item_dir, exist_ok=True)

            root_items = api_items(public_link, "", session)
            if not root_items:
                fail_count += 1
                failed_articles.append(article)
                print(f"  -> FAIL: не удалось прочитать ссылку: {public_link}")

                report_rows.append({
                    "Артикул": article,
                    "Наименование": name,
                    "Ссылка": public_link,
                    "Статус": "FAIL",
                    "Превью": "нет",
                    "Галерея": 0,
                    "Источник": "",
                    "Примечание": "не удалось прочитать содержимое ссылки",
                })
                continue

            preview_file = next(
                (
                    item for item in root_items
                    if item.get("type") == "file" and is_preview_image(item.get("name", ""))
                ),
                None
            )

            images_dir_path, images_dir_note = find_images_dir(public_link, session)

            if images_dir_path:
                image_items = api_items(public_link, images_dir_path, session)
                gallery_files = sorted(
                    [
                        item for item in image_items
                        if item.get("type") == "file" and is_gallery_image(item.get("name", ""))
                    ],
                    key=lambda x: x.get("name", "")
                )
                gallery_source = "папка Images"
            else:
                gallery_files = sorted(
                    [
                        item for item in root_items
                        if item.get("type") == "file" and is_any_image(item.get("name", ""))
                    ],
                    key=lambda x: x.get("name", "")
                )
                gallery_source = "корень ссылки"

            if not preview_file and gallery_files:
                preview_file = gallery_files[0]

            saved_preview = False
            saved_gallery = 0
            fail_reason = ""

            if preview_file and preview_file.get("file"):
                preview_path = os.path.join(item_dir, "preview.webp")
                ok, reason = download_and_compress_image(preview_file["file"], preview_path, session)
                saved_preview = ok
                if not ok:
                    fail_reason = f"превью не скачалось: {reason}"
            else:
                fail_reason = "превью не найдено"

            for file_item in gallery_files:
                if saved_gallery >= MAX_GALLERY:
                    break

                file_url = file_item.get("file")
                if not file_url:
                    continue

                gallery_path = os.path.join(item_dir, f"gallery_{saved_gallery + 1:02}.webp")
                ok, _ = download_and_compress_image(file_url, gallery_path, session)
                if ok:
                    saved_gallery += 1

            if saved_preview or saved_gallery > 0:
                ok_count += 1

                if saved_gallery < MAX_GALLERY:
                    partial_count += 1
                    partial_articles.append(article)
                    status = "PARTIAL"
                    note = f"{images_dir_note}; скачано меньше {MAX_GALLERY} файлов галереи"
                    print(f"  -> PARTIAL: превью={'да' if saved_preview else 'нет'}, галерея={saved_gallery}")
                else:
                    status = "OK"
                    note = images_dir_note
                    print(f"  -> OK: превью={'да' if saved_preview else 'нет'}, галерея={saved_gallery}")
            else:
                fail_count += 1
                failed_articles.append(article)
                status = "FAIL"
                note = fail_reason or images_dir_note
                print("  -> FAIL: не удалось скачать ни превью, ни галерею")

            report_rows.append({
                "Артикул": article,
                "Наименование": name,
                "Ссылка": public_link,
                "Статус": status,
                "Превью": "да" if saved_preview else "нет",
                "Галерея": saved_gallery,
                "Источник": gallery_source,
                "Примечание": note,
            })

            time.sleep(SLEEP_BETWEEN_ITEMS)

    report_df = pd.DataFrame(report_rows)
    without_images_df = report_df[report_df["Статус"] == "FAIL"].copy()

    with pd.ExcelWriter(REPORT_FILE, engine="openpyxl") as writer:
        report_df.to_excel(writer, sheet_name="report", index=False)
        without_images_df.to_excel(writer, sheet_name="without_images", index=False)

    print("\n===== ИТОГ =====")
    print(f"Успешно: {ok_count}")
    print(f"Частично: {partial_count}")
    print(f"Ошибок: {fail_count}")
    print(f"Пропусков: {skip_count}")
    print(f"Отчёт сохранён: {REPORT_FILE}")

    print("\n===== НЕ УДАЛОСЬ СКАЧАТЬ =====")
    if failed_articles:
        for article in failed_articles:
            print(f" - {article}")
    else:
        print("Нет")

    print("\n===== ЧАСТИЧНОЕ ЗАПОЛНЕНИЕ =====")
    if partial_articles:
        for article in partial_articles:
            print(f" - {article}")
    else:
        print("Нет")


if __name__ == "__main__":
    main()